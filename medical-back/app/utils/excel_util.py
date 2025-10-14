# Excel 工具
# 迁移自 py_tools.utils.excel_util.py
import openpyxl
from typing import List

def read_excel(file_path: str) -> List[list]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]

def write_excel(file_path: str, data: List[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(file_path)